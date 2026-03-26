# ACTIVE: maintained canon/theory asset rebuild entrypoint still referenced by current planning docs.
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from functools import lru_cache
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from juthoor_arabicgenome_lv1.core.feature_decomposition import (
    decompose_semantic_text,
    feature_categories,
)
from juthoor_arabicgenome_lv1.factory.scoring import (
    build_consensus_scholar_letters,
    build_nucleus_score_rows,
    invert_features_extended,
)
from juthoor_arabicgenome_lv1.factory.root_predictor import (
    build_root_prediction_rows,
    build_root_prediction_rows_all_scholars,
    summarize_root_predictions,
)
from juthoor_arabicgenome_lv1.factory.cross_lingual_projection import (
    build_non_semitic_projection_rows,
    build_semitic_projection_rows,
    load_benchmark_rows,
    non_semitic_projection_summary,
    projection_summary,
)
from juthoor_arabicgenome_lv1.factory.cross_lingual_scoring import (
    score_projection_row,
    summarize_projection_scores,
)
from juthoor_arabicgenome_lv1.factory.independent_letter_derivation import (
    derive_independent_letter_meanings,
    render_independent_letter_genome_markdown,
)
from juthoor_arabicgenome_lv1.factory.third_letter_profiles import (
    build_third_letter_modifier_profiles,
    render_third_letter_modifier_profiles_markdown,
)


LV1_ROOT = Path(__file__).resolve().parents[2]
REPO_ROOT = LV1_ROOT.parent
DOC_ROOT = LV1_ROOT / "docs" / "The Arabic Tongue (nature-genome-application)"
THEORY_ROOT = DOC_ROOT / "Languistic theories"
SOURCE_DOC_ROOT = REPO_ROOT.parent / "The Arabic Tongue (nature-genome-application)"
if not SOURCE_DOC_ROOT.exists():
    SOURCE_DOC_ROOT = DOC_ROOT
SOURCE_THEORY_ROOT = SOURCE_DOC_ROOT / "Languistic theories"
SOURCE_MUAJAM_XLSX = SOURCE_DOC_ROOT / "Muajam Ishtiqaqi" / "المعجم_الاشتقاقي_Juthoor_v2.xlsx"
MUAJAM_ROOT = LV1_ROOT / "data" / "muajam"
CANON_ROOT = LV1_ROOT / "data" / "theory_canon"
LETTERS_OUT = CANON_ROOT / "letters"
BINARY_OUT = CANON_ROOT / "binary_fields"
ROOTS_OUT = CANON_ROOT / "roots"
REGISTRIES_OUT = CANON_ROOT / "registries"
OUTPUT_ROOT = REPO_ROOT / "outputs" / "lv1_scoring"
LV2_BENCHMARK = REPO_ROOT / "Juthoor-CognateDiscovery-LV2" / "resources" / "benchmarks" / "cognate_gold.jsonl"

LETTER_NAME_TO_CHAR = {
    "الهمزة": "ء",
    "ألف المد": "ا",
    "ألف": "ا",
    "باء": "ب",
    "تاء": "ت",
    "ثاء": "ث",
    "جيم": "ج",
    "حاء": "ح",
    "خاء": "خ",
    "دال": "د",
    "ذال": "ذ",
    "راء": "ر",
    "زاء": "ز",
    "زاي": "ز",
    "سين": "س",
    "شين": "ش",
    "صاد": "ص",
    "ضاد": "ض",
    "طاء": "ط",
    "ظاء": "ظ",
    "عين": "ع",
    "غين": "غ",
    "فاء": "ف",
    "قاف": "ق",
    "كاف": "ك",
    "لام": "ل",
    "ميم": "م",
    "نون": "ن",
    "هاء": "هـ",
    "واو": "و",
    "ياء": "ي",
}

CHAR_TO_LETTER_NAME = {
    "ء": "الهمزة",
    "ا": "ألف المد",
    "ب": "الباء",
    "ت": "التاء",
    "ث": "الثاء",
    "ج": "الجيم",
    "ح": "الحاء",
    "خ": "الخاء",
    "د": "الدال",
    "ذ": "الذال",
    "ر": "الراء",
    "ز": "الزاي",
    "س": "السين",
    "ش": "الشين",
    "ص": "الصاد",
    "ض": "الضاد",
    "ط": "الطاء",
    "ظ": "الظاء",
    "ع": "العين",
    "غ": "الغين",
    "ف": "الفاء",
    "ق": "القاف",
    "ك": "الكاف",
    "ل": "اللام",
    "م": "الميم",
    "ن": "النون",
    "هـ": "الهاء",
    "و": "الواو",
    "ي": "الياء",
}
CORE_REGISTRY_LETTERS = frozenset(CHAR_TO_LETTER_NAME.keys())

NEILI_LETTERS = frozenset({"د", "ح", "ر", "ت", "ك", "م", "ب", "ع", "ل", "ي"})
JABAL_RAW_DESCRIPTION_OVERRIDES = {
    "ب": "ظهور وخروج",
    "م": "تجمع وتلاصق",
    "ع": "ظهور وعمق",
    "غ": "باطن واشتمال",
}
JABAL_FEATURE_OVERRIDES = {
    "ب": ("ظهور", "خروج"),
    "م": ("تجمع", "تلاصق"),
    "ع": ("ظهور", "عمق"),
    "غ": ("باطن", "اشتمال"),
}
ASIM_FEATURE_OVERRIDES = {
    "ا": ("اتصال", "وحدة"),
    "س": ("امتداد", "ظهور"),
    "ك": ("تجمع", "احتواء"),
}
ABBAS_FALLBACK_DETAILS = {
    "هـ": {
        "raw_description": "اهتزاز واضطراب وفراغ",
        "sensory_category": "شعورية (حلقية)",
        "mechanism_type": "إيحائية",
    }
}


def _read_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def _write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


@lru_cache(maxsize=1)
def _source_workbook():
    return load_workbook(SOURCE_MUAJAM_XLSX, read_only=True, data_only=True)


def _normalize_compact_text(value: Any) -> str:
    return re.sub(r"\s+", "", str(value or ""))


def _normalize_letter_symbol(symbol: str | None) -> str:
    value = _normalize_compact_text(symbol)
    if value in {"ه", "هة"}:
        return "هـ"
    if value in {"أ", "إ", "آ"}:
        return "ء"
    if value == "ى":
        return "ي"
    return value


def _extract_last_arabic_letter(text: str) -> str:
    matches = re.findall(r"[ءاأإآبتثجحخدذرزسشصضطظعغفقكلمنهويى]", text)
    if not matches:
        return ""
    last = matches[-1]
    if last == "ه":
        return "هـ"
    if last in {"أ", "إ", "آ"}:
        return "ء"
    if last == "ى":
        return "ي"
    return last


def _iter_jabal_xlsx_rows() -> list[tuple[Any, ...]]:
    ws = _source_workbook()["المعجم الكامل"]
    return [row for row in ws.iter_rows(min_row=2, values_only=True) if any(row)]


def _letter_row(
    *,
    letter: str,
    scholar: str,
    raw_description: str,
    source_document: str,
    letter_name: str | None = None,
    sensory_category: str | None = None,
    kinetic_gloss: str | None = None,
    confidence: str = "medium",
) -> dict[str, Any]:
    features = decompose_semantic_text(raw_description)
    return {
        "letter": letter,
        "letter_name": letter_name,
        "scholar": scholar,
        "raw_description": raw_description,
        "atomic_features": list(features),
        "feature_categories": list(feature_categories(features)),
        "sensory_category": sensory_category,
        "kinetic_gloss": kinetic_gloss,
        "source_document": source_document,
        "confidence": confidence,
    }


def _load_jabal_letters() -> list[dict[str, Any]]:
    ws = _source_workbook()["معاني الحروف"]
    out: list[dict[str, Any]] = []
    for letter_name, letter, meaning in ws.iter_rows(min_row=2, values_only=True):
        if not letter:
            continue
        normalized_letter = _normalize_letter_symbol(str(letter).strip())
        raw_description = JABAL_RAW_DESCRIPTION_OVERRIDES.get(
            normalized_letter,
            str(meaning or "").strip(),
        )
        payload = _letter_row(
            letter=normalized_letter,
            letter_name=str(letter_name).strip(),
            scholar="jabal",
            raw_description=raw_description,
            source_document=f"{SOURCE_MUAJAM_XLSX}#معاني الحروف",
            confidence="high",
        )
        if normalized_letter in JABAL_FEATURE_OVERRIDES:
            features = JABAL_FEATURE_OVERRIDES[normalized_letter]
            payload["atomic_features"] = list(features)
            payload["feature_categories"] = list(feature_categories(features))
            payload["source_note"] = "Yassin-confirmed 2026-03-26 correction"
        out.append(payload)
    return out


def _extract_markdown_table(text: str, start_marker: str) -> list[list[str]]:
    lines = text.splitlines()
    capture = False
    rows: list[list[str]] = []
    for line in lines:
        if start_marker in line:
            capture = True
            continue
        if capture and line.startswith("|"):
            cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
            if set(cells[0]) == {"-"}:
                continue
            rows.append(cells)
        elif capture and rows:
            break
    return rows


def _load_neili_letters() -> list[dict[str, Any]]:
    summary = (SOURCE_DOC_ROOT / "ملخص_الدلالة_الصوتية_العربية.md").read_text(encoding="utf-8")
    rows = _extract_markdown_table(summary, "### الحروف العشرة المستكشفة")
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        letter = _normalize_letter_symbol(row[0].replace("*", "").replace(" ", ""))
        kinetic = row[1]
        gloss = row[2]
        out.append(
            _letter_row(
                letter=letter,
                scholar="neili",
                raw_description=gloss,
                kinetic_gloss=kinetic,
                source_document=str(SOURCE_DOC_ROOT / "ملخص_الدلالة_الصوتية_العربية.md"),
                confidence="high",
            )
        )
    return out


def _load_abbas_letters() -> list[dict[str, Any]]:
    source_path = SOURCE_DOC_ROOT / "ABBAS_LETTER_CLASSIFICATION.md"
    text = source_path.read_text(encoding="utf-8")

    detail_map: dict[str, dict[str, str]] = {}
    for row in _extract_markdown_table(text, "### A. الحروف الهيجانية"):
        if len(row) < 4 or row[0] == "Letter":
            continue
        letter = re.search(r"\*\*(.+?)\*\*", row[0])
        if not letter:
            continue
        detail_map[_normalize_letter_symbol(letter.group(1))] = {
            "sensory_category": row[1].strip(),
            "raw_description": row[2].strip(),
            "mechanism_type": "هيجانية",
            "source_certainty": "يقيناً",
            "note": row[3].strip(),
        }
    for row in _extract_markdown_table(text, "### B. الحروف الإيمائية"):
        if len(row) < 4 or row[0] == "Letter":
            continue
        letter = re.search(r"\*\*(.+?)\*\*", row[0])
        if not letter:
            continue
        detail_map[_normalize_letter_symbol(letter.group(1))] = {
            "sensory_category": row[1].strip(),
            "raw_description": row[2].strip(),
            "mechanism_type": "إيمائية",
            "source_certainty": "ترجيح شديد",
            "kinetic_gloss": row[3].strip(),
        }
    for row in _extract_markdown_table(text, "### C. الحروف الإيحائية"):
        if len(row) < 3 or row[0] == "Letter":
            continue
        letter = re.search(r"\*\*(.+?)\*\*", row[0])
        if not letter:
            continue
        detail_map[_normalize_letter_symbol(letter.group(1))] = {
            "sensory_category": row[1].strip(),
            "raw_description": row[2].strip(),
            "mechanism_type": "إيحائية",
            "source_certainty": "reviewed",
        }

    out: list[dict[str, Any]] = []
    for row in _extract_markdown_table(text, "## Summary Table — All 28 Letters"):
        if len(row) < 5 or row[0] == "#" or row[0].startswith("---"):
            continue
        _, letter, sensory_category, mechanism_type, source_certainty = row
        letter = _normalize_letter_symbol(letter)
        detail = detail_map.get(letter, {})
        if not detail and letter in ABBAS_FALLBACK_DETAILS:
            detail = ABBAS_FALLBACK_DETAILS[letter]
        payload = _letter_row(
            letter=letter,
            letter_name=CHAR_TO_LETTER_NAME.get(letter),
            scholar="hassan_abbas",
            raw_description=detail.get("raw_description") or sensory_category,
            sensory_category=detail.get("sensory_category") or sensory_category,
            kinetic_gloss=detail.get("kinetic_gloss"),
            source_document=str(source_path),
            confidence="high" if source_certainty.strip() == "يقيناً" else "medium",
        )
        payload["mechanism_type"] = detail.get("mechanism_type") or mechanism_type.strip()
        payload["source_certainty"] = source_certainty.strip()
        if detail.get("note"):
            payload["note"] = detail["note"]
        out.append(payload)
    return out


def _extract_asim_full_table_rows(text: str) -> list[list[str]]:
    rows: list[list[str]] = []
    for line in text.splitlines():
        if not line.startswith("|"):
            continue
        cells = [cell.strip() for cell in line.strip().strip("|").split("|")]
        if len(cells) != 4:
            continue
        if cells[0] in {"الترتيب", "---"}:
            continue
        if not re.fullmatch(r"[١٢٣٤٥٦٧٨٩٠]+", cells[0]):
            continue
        rows.append(cells)
    return rows


def _normalize_letter_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip())


def _load_asim_letters() -> list[dict[str, Any]]:
    source_path = SOURCE_DOC_ROOT / "LV1_VERIFIED_DATA_AUDIT.md"
    rows = _extract_markdown_table(source_path.read_text(encoding="utf-8"), "**عاصم's complete 28-letter table:**")
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if len(row) < 3:
            continue
        _, raw_name, raw_gloss = row[:3]
        letter_name = _normalize_letter_name(raw_name)
        letter = LETTER_NAME_TO_CHAR.get(letter_name)
        if not letter:
            continue
        payload = _letter_row(
            letter=letter,
            letter_name=letter_name,
            scholar="asim_al_masri",
            raw_description=raw_gloss.strip(),
            kinetic_gloss=raw_gloss.strip(),
            source_document=f"{source_path}#asim-complete-table",
        )
        if not payload["atomic_features"] and letter in ASIM_FEATURE_OVERRIDES:
            features = ASIM_FEATURE_OVERRIDES[letter]
            payload["atomic_features"] = list(features)
            payload["feature_categories"] = list(feature_categories(features))
        payload["continues_neili"] = letter in NEILI_LETTERS
        out.append(payload)
    return out


def _load_anbar_letters() -> list[dict[str, Any]]:
    source_path = SOURCE_DOC_ROOT / "LV1_VERIFIED_DATA_AUDIT.md"
    rows = _extract_markdown_table(
        source_path.read_text(encoding="utf-8"),
        "**عنبر's complete extracted letter meanings:**",
    )
    out: list[dict[str, Any]] = []
    for row in rows[1:]:
        if len(row) < 3:
            continue
        raw_letter, phonetic_group, gloss = row[:3]
        letter = raw_letter.replace("*", "").strip()
        is_contextual = "سياقية" in phonetic_group or "⚠" in gloss
        normalized_letter = _normalize_letter_symbol(letter) if len(letter) == 1 else letter
        payload = _letter_row(
            letter=normalized_letter,
            letter_name=CHAR_TO_LETTER_NAME.get(normalized_letter, letter),
            scholar="anbar",
            raw_description=gloss.replace("⚠️", "").replace("⚠", "").strip(),
            kinetic_gloss=None,
            source_document=f"{source_path}#anbar-complete-table",
            confidence="medium" if is_contextual else "high",
        )
        payload["phonetic_group"] = phonetic_group.replace("**", "").strip()
        if is_contextual:
            payload["derivation"] = "contextual"
        out.append(payload)
    return out


def _build_jabal_nuclei() -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in _iter_jabal_xlsx_rows():
        key = _normalize_compact_text(row[2])
        letter1 = _normalize_compact_text(row[3])
        letter2 = _normalize_compact_text(row[5])
        tri_root = _normalize_compact_text(row[8])
        shared_meaning = str(row[7] or "").strip()
        grouped.setdefault(
            key,
            {
                "binary_root": key,
                "letter1": letter1,
                "letter2": letter2,
                "jabal_shared_meaning": shared_meaning,
                "jabal_features": list(decompose_semantic_text(shared_meaning)),
                "member_roots": [],
                "member_count": 0,
                "reverse_exists": False,
                "reverse_root": None,
                "model_a_score": None,
                "model_b_score": None,
                "model_c_score": None,
                "model_d_score": None,
                "best_model": None,
                "golden_rule_score": None,
                "status": "empty",
                "bab": str(row[0] or ""),
                "source": str(SOURCE_MUAJAM_XLSX),
            },
        )
        if tri_root and tri_root not in grouped[key]["member_roots"]:
            grouped[key]["member_roots"].append(tri_root)
    keys = set(grouped)
    for key, payload in grouped.items():
        payload["member_count"] = len(payload["member_roots"])
        payload["reverse_exists"] = key[::-1] in keys and key[::-1] != key
        payload["reverse_root"] = key[::-1] if payload["reverse_exists"] else None
    return sorted(grouped.values(), key=lambda item: item["binary_root"])


def _build_jabal_roots() -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for row in _iter_jabal_xlsx_rows():
        root = _normalize_compact_text(row[8])
        third_letter = _extract_last_arabic_letter(root)
        axial_meaning = str(row[10] or "").strip()
        quranic_verse = str(row[11] or "").strip() or None
        out.append(
            {
                "root": root,
                "binary_nucleus": _normalize_compact_text(row[2]),
                "third_letter": third_letter,
                "jabal_axial_meaning": axial_meaning,
                "jabal_features": list(decompose_semantic_text(axial_meaning)),
                "predicted_meaning": None,
                "predicted_features": None,
                "prediction_score": None,
                "quranic_verse": quranic_verse,
                "is_quranic": bool(quranic_verse),
                "bab": str(row[0] or ""),
                "source": str(SOURCE_MUAJAM_XLSX),
                "status": "empty",
            }
        )
    return out


def _scholar_letter_map(rows: list[dict[str, Any]]) -> dict[str, dict[str, dict[str, Any]]]:
    grouped: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for row in rows:
        grouped[row["scholar"]][row["letter"]] = row
    return dict(grouped)


def _consensus_rows_from_map(
    scholar_map: dict[str, dict[str, dict[str, Any]]],
    *,
    scholar_name: str,
    mode: str,
) -> list[dict[str, Any]]:
    consensus_map = build_consensus_scholar_letters(scholar_map, mode=mode)
    rows: list[dict[str, Any]] = []
    for letter, payload in sorted(consensus_map.items()):
        if letter not in CORE_REGISTRY_LETTERS:
            continue
        rows.append(
            {
                "letter": letter,
                "letter_name": payload.get("letter_name"),
                "scholar": scholar_name,
                "raw_description": " / ".join(payload.get("atomic_features") or ()) or None,
                "atomic_features": payload.get("atomic_features") or (),
                "source_document": f"generated:{scholar_name}",
                "articulatory_features": payload.get("articulatory_features"),
                "confidence": "generated",
                "mode": mode,
                "support_counts": payload.get("support_counts"),
            }
        )
    return rows


def _confidence_tier(agreement_level: str, source_count: int) -> str:
    if agreement_level == "consensus":
        return "high"
    if agreement_level == "majority":
        return "medium"
    if source_count >= 2:
        return "low"
    return "stub"


def _build_letter_registry_rows(scholar_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    core_rows = [
        row
        for row in scholar_rows
        if row.get("letter") in CORE_REGISTRY_LETTERS
    ]
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in core_rows:
        grouped[row["letter"]].append(row)

    registry_rows: list[dict[str, Any]] = []
    for letter, rows in sorted(grouped.items()):
        feature_counts: defaultdict[str, set[str]] = defaultdict(set)
        for row in rows:
            for feature in row.get("atomic_features") or []:
                feature_counts[feature].add(row["scholar"])
        consensus = sorted(feature for feature, scholars in feature_counts.items() if len(scholars) >= 2)
        scholar_count = len({row["scholar"] for row in rows})
        if scholar_count <= 1:
            agreement_level = "unknown"
        elif scholar_count >= 3 and consensus:
            agreement_level = "consensus"
        elif consensus:
            agreement_level = "majority"
        else:
            agreement_level = "contested"

        jabal_row = next((row for row in rows if row["scholar"] == "jabal"), None)
        asim_row = next((row for row in rows if row["scholar"] == "asim_al_masri"), None)
        abbas_row = next((row for row in rows if row["scholar"] == "hassan_abbas"), None)

        source_entries = []
        for row in rows:
            claim_type = "semantic"
            if row["scholar"] == "asim_al_masri":
                claim_type = "kinetic"
            elif row["scholar"] == "hassan_abbas":
                claim_type = "sensory"
            source_entries.append(
                {
                    "source_id": f"{row['scholar']}-{letter}",
                    "scholar": row["scholar"],
                    "claim_type": claim_type,
                    "claim_text": row["raw_description"],
                    "document_ref": row["source_document"],
                    "curation_status": "reviewed",
                }
            )

        registry_rows.append(
            {
                "letter": letter,
                "letter_name": (jabal_row or rows[0]).get("letter_name"),
                "canonical_semantic_gloss": (jabal_row or rows[0]).get("raw_description"),
                "canonical_kinetic_gloss": (asim_row or {}).get("kinetic_gloss"),
                "canonical_sensory_gloss": (abbas_row or {}).get("sensory_category"),
                "articulatory_features": None,
                "sources": source_entries,
                "agreement_level": agreement_level,
                "confidence_tier": _confidence_tier(agreement_level, scholar_count),
                "status": "draft",
            }
        )
    return registry_rows


def _build_golden_rule_report(nuclei_rows: list[dict[str, Any]]) -> dict[str, Any]:
    lookup = {row["binary_root"]: row for row in nuclei_rows}
    seen: set[str] = set()
    pair_rows: list[dict[str, Any]] = []
    confirmed = 0
    for root, row in lookup.items():
        reverse = root[::-1]
        if reverse not in lookup or root == reverse or reverse in seen:
            continue
        seen.add(root)
        seen.add(reverse)
        forward_features = tuple(row["jabal_features"])
        reverse_features = tuple(lookup[reverse]["jabal_features"])
        inverted_forward = set(invert_features_extended(forward_features))
        overlap = sorted(inverted_forward & set(reverse_features))
        is_confirmed = bool(overlap)
        if is_confirmed:
            confirmed += 1
        pair_rows.append(
            {
                "forward": root,
                "reverse": reverse,
                "forward_features": list(forward_features),
                "reverse_features": list(reverse_features),
                "inverted_forward_features": list(inverted_forward),
                "overlap_with_reverse": overlap,
                "confirmed": is_confirmed,
            }
        )
    total = len(pair_rows)
    return {
        "reversible_pairs": total,
        "confirmed_pairs": confirmed,
        "confirmation_rate": round(confirmed / total, 6) if total else 0.0,
        "pairs": pair_rows,
    }


def _build_full_letter_nucleus_evidence(nuclei_rows: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    out: dict[str, dict[str, Any]] = {}
    for row in nuclei_rows:
        letters = re.findall(r"[ءاأإآبتثجحخدذرزسشصضطظعغفقكلمنهويى]", row["binary_root"])
        if len(letters) != 2:
            continue
        letter1 = _normalize_letter_symbol(letters[0])
        letter2 = _normalize_letter_symbol(letters[1])
        entry = {
            "nucleus": row["binary_root"],
            "shared_meaning": row["jabal_shared_meaning"],
            "features": list(row["jabal_features"]),
            "member_count": int(row["member_count"]),
        }
        out.setdefault(
            letter1,
            {"letter": letter1, "as_letter1": [], "as_letter2": [], "count_l1": 0, "count_l2": 0, "total": 0},
        )
        out.setdefault(
            letter2,
            {"letter": letter2, "as_letter1": [], "as_letter2": [], "count_l1": 0, "count_l2": 0, "total": 0},
        )
        out[letter1]["as_letter1"].append(entry)
        out[letter2]["as_letter2"].append(entry)
    for payload in out.values():
        payload["count_l1"] = len(payload["as_letter1"])
        payload["count_l2"] = len(payload["as_letter2"])
        payload["total"] = payload["count_l1"] + payload["count_l2"]
    return dict(sorted(out.items()))


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")

    base_scholar_rows = (
        _load_jabal_letters()
        + _load_neili_letters()
        + _load_abbas_letters()
        + _load_asim_letters()
        + _load_anbar_letters()
    )
    nuclei_rows = _build_jabal_nuclei()
    root_rows = _build_jabal_roots()
    base_scholar_map = _scholar_letter_map(base_scholar_rows)
    consensus_strict_rows = _consensus_rows_from_map(
        base_scholar_map,
        scholar_name="consensus_strict",
        mode="strict",
    )
    consensus_weighted_rows = _consensus_rows_from_map(
        base_scholar_map,
        scholar_name="consensus_weighted",
        mode="weighted",
    )
    scoring_scholar_rows = base_scholar_rows + consensus_strict_rows + consensus_weighted_rows
    scholar_map = _scholar_letter_map(scoring_scholar_rows)
    score_rows = build_nucleus_score_rows(nuclei_rows, scholar_map)
    golden_rule = _build_golden_rule_report(nuclei_rows)
    full_letter_nucleus_evidence = _build_full_letter_nucleus_evidence(nuclei_rows)
    independent_letter_derivations = derive_independent_letter_meanings(
        full_letter_nucleus_evidence,
        base_scholar_rows,
    )
    independent_letter_genome_md = render_independent_letter_genome_markdown(independent_letter_derivations)
    root_prediction_rows = build_root_prediction_rows_all_scholars(
        root_rows,
        nuclei_rows,
        scholar_map,
        scholars=(
            "jabal",
            "asim_al_masri",
            "hassan_abbas",
            "consensus_strict",
            "consensus_weighted",
        ),
    )
    third_letter_profiles = build_third_letter_modifier_profiles(root_prediction_rows)
    third_letter_profiles_md = render_third_letter_modifier_profiles_markdown(third_letter_profiles)
    root_score_matrix = summarize_root_predictions(root_prediction_rows)
    jabal_root_prediction_rows = [row for row in root_prediction_rows if row["scholar"] == "jabal"]
    benchmark_rows = load_benchmark_rows(LV2_BENCHMARK)
    semitic_projection_rows = build_semitic_projection_rows(jabal_root_prediction_rows, benchmark_rows)
    semitic_projection_summary = projection_summary(semitic_projection_rows, benchmark_rows)
    semitic_scored_rows = [score_projection_row(row) for row in semitic_projection_rows]
    semitic_scoring_summary = summarize_projection_scores(semitic_scored_rows)
    non_semitic_projection_rows = build_non_semitic_projection_rows(jabal_root_prediction_rows, benchmark_rows)
    non_semitic_projection_summary_payload = non_semitic_projection_summary(
        non_semitic_projection_rows,
        benchmark_rows,
    )
    non_semitic_scored_rows = [score_projection_row(row) for row in non_semitic_projection_rows]
    non_semitic_scoring_summary = summarize_projection_scores(non_semitic_scored_rows)

    prediction_by_root = {row["root"]: row for row in jabal_root_prediction_rows}
    for row in root_rows:
        prediction = prediction_by_root.get(row["root"])
        if not prediction:
            continue
        row["predicted_meaning"] = prediction["predicted_meaning"]
        row["predicted_features"] = prediction["predicted_features"]
        row["prediction_score"] = prediction["weighted_jaccard"]
        row["status"] = "predicted" if prediction["predicted_features"] else "empty"

    by_scholar: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in scoring_scholar_rows:
        by_scholar[row["scholar"]].append(row)
    letter_registry_rows = _build_letter_registry_rows(base_scholar_rows)
    for scholar, rows in by_scholar.items():
        _write_jsonl(LETTERS_OUT / f"{scholar}_letters.jsonl", sorted(rows, key=lambda item: item["letter"]))
    _write_jsonl(BINARY_OUT / "jabal_nuclei_raw.jsonl", nuclei_rows)
    _write_jsonl(ROOTS_OUT / "jabal_roots_raw.jsonl", root_rows)
    _write_jsonl(REGISTRIES_OUT / "letters.jsonl", letter_registry_rows)
    _write_json(OUTPUT_ROOT / "nucleus_score_matrix.json", score_rows)
    _write_json(OUTPUT_ROOT / "golden_rule_report.json", golden_rule)
    _write_json(OUTPUT_ROOT / "full_letter_nucleus_evidence.json", full_letter_nucleus_evidence)
    _write_json(OUTPUT_ROOT / "independent_letter_derivations.json", independent_letter_derivations)
    _write_text(OUTPUT_ROOT / "INDEPENDENT_ARABIC_LETTER_GENOME.md", independent_letter_genome_md)
    _write_json(OUTPUT_ROOT / "root_predictions.json", root_prediction_rows)
    _write_json(OUTPUT_ROOT / "third_letter_modifier_profiles.json", third_letter_profiles)
    _write_text(OUTPUT_ROOT / "THIRD_LETTER_MODIFIER_PROFILES.md", third_letter_profiles_md)
    _write_json(OUTPUT_ROOT / "root_score_matrix.json", root_score_matrix)
    _write_json(OUTPUT_ROOT / "benchmark_semitic_projections.json", semitic_projection_rows)
    _write_json(OUTPUT_ROOT / "benchmark_semitic_projection_summary.json", semitic_projection_summary)
    _write_json(OUTPUT_ROOT / "benchmark_semitic_scored_projections.json", semitic_scored_rows)
    _write_json(OUTPUT_ROOT / "benchmark_semitic_scoring_summary.json", semitic_scoring_summary)
    _write_json(OUTPUT_ROOT / "benchmark_non_semitic_projections.json", non_semitic_projection_rows)
    _write_json(OUTPUT_ROOT / "benchmark_non_semitic_projection_summary.json", non_semitic_projection_summary_payload)
    _write_json(OUTPUT_ROOT / "benchmark_non_semitic_scored_projections.json", non_semitic_scored_rows)
    _write_json(OUTPUT_ROOT / "benchmark_non_semitic_scoring_summary.json", non_semitic_scoring_summary)

    summary = {
        "scholars": {scholar: len(rows) for scholar, rows in by_scholar.items()},
        "nuclei": len(nuclei_rows),
        "roots": len(root_rows),
        "score_rows": len(score_rows),
        "root_prediction_rows": len(root_prediction_rows),
        "golden_rule_pairs": golden_rule["reversible_pairs"],
        "benchmark_semitic_rows": len(semitic_projection_rows),
        "benchmark_non_semitic_rows": len(non_semitic_projection_rows),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
