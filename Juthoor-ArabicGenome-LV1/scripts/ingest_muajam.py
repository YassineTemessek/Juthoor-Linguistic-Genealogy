"""
ingest_muajam.py — Phase 2 ingest for the Muajam Ishtiqaqi XLSX data.

Reads:
  - 24 per-letter XLSX files in Muajam Ishtiqaqi/Tables_Juthoor/
  - Master file: Muajam Ishtiqaqi/المعجم_الاشتقاقي_Juthoor_v2.xlsx (sheet: معاني الحروف)

Writes:
  - data/muajam/letter_meanings.jsonl
  - data/muajam/roots.jsonl
"""

import sys
import json
import re
from pathlib import Path

import openpyxl


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_spaces(text: str) -> str:
    """Remove all whitespace from a string (collapses spaced root to compact)."""
    if not text:
        return ""
    return re.sub(r"\s+", "", text)


def cell_str(cell) -> str:
    """Return cell value as a stripped string, or empty string for None."""
    val = cell.value
    if val is None:
        return ""
    return str(val).strip()


def extract_single_arabic_char(text: str) -> str:
    """
    Extract the first Arabic letter character from a string.
    Arabic block: U+0600–U+06FF  (covers letters, diacritics, etc.)
    We want the first base letter (not diacritics).
    Arabic base letters sit in ~U+0621–U+064A.
    """
    for ch in text:
        cp = ord(ch)
        if 0x0621 <= cp <= 0x064A:
            return ch
    return ""


# ---------------------------------------------------------------------------
# Step 1: Extract letter meanings from the master XLSX
# ---------------------------------------------------------------------------

def extract_letter_meanings(master_path: Path) -> list[dict]:
    """Parse the معاني الحروف sheet from the master XLSX."""
    wb = openpyxl.load_workbook(master_path, read_only=True, data_only=True)

    sheet_name = "معاني الحروف"
    if sheet_name not in wb.sheetnames:
        available = wb.sheetnames
        wb.close()
        raise ValueError(
            f"Sheet '{sheet_name}' not found in master XLSX. "
            f"Available sheets: {available}"
        )

    ws = wb[sheet_name]
    records = []
    first_row = True

    for row in ws.iter_rows():
        if first_row:
            first_row = False
            continue  # skip header

        letter_name = cell_str(row[0]) if len(row) > 0 else ""
        letter_sym  = cell_str(row[1]) if len(row) > 1 else ""
        meaning     = cell_str(row[2]) if len(row) > 2 else ""

        # Skip empty rows
        if not letter_name and not letter_sym:
            continue

        records.append({
            "letter": letter_sym,
            "letter_name": letter_name,
            "meaning": meaning,
        })

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Step 2: Extract roots from a single per-letter XLSX file
# ---------------------------------------------------------------------------

def extract_roots_from_file(xlsx_path: Path) -> list[dict]:
    """
    Parse one per-letter XLSX file and return a list of root entry dicts.

    Column layout (0-indexed):
      0  الباب              — BAB letter name  (e.g. "الباء")
      1  معنى الباب         — BAB meaning
      2  الجذر الثنائي      — biconsonantal root spaced (e.g. "ب ب")
      3  ح١                — letter 1
      4  معنى ح١            — letter 1 meaning
      5  ح٢                — letter 2
      6  معنى ح٢            — letter 2 meaning
      7  المعنى المشترك…    — shared / biconsonantal root meaning
      8  التركيب (الجذر الثلاثي) — triconsonantal root spaced (e.g. "ب و ب")
      9  الحرف المُضاف      — added letter
      10 المعنى المحوري    — axial meaning
      11 التطبيق القرآني   — Quranic example
    """
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active  # single sheet per file

    records = []
    first_row = True

    for row in ws.iter_rows():
        if first_row:
            first_row = False
            continue  # skip header

        def col(idx):
            return cell_str(row[idx]) if len(row) > idx else ""

        bab_name       = col(0)
        bab_meaning    = col(1)
        bi_root_spaced = col(2)
        letter1        = col(3)
        letter1_meaning= col(4)
        letter2        = col(5)
        letter2_meaning= col(6)
        bi_meaning     = col(7)
        tri_root_spaced= col(8)
        added_letter   = col(9)
        axial_meaning  = col(10)
        quran_example  = col(11)

        # Skip rows that have no meaningful content
        if not bab_name and not bi_root_spaced and not tri_root_spaced:
            continue

        # Derive compact forms
        bi_root  = strip_spaces(bi_root_spaced)
        tri_root = strip_spaces(tri_root_spaced)

        # Extract single-char BAB identifier from letter1 column (most reliable)
        bab_char = letter1 if len(letter1) == 1 else extract_single_arabic_char(bab_name)

        records.append({
            "bab": bab_char,
            "bab_name": bab_name,
            "bab_meaning": bab_meaning,
            "binary_root": bi_root,
            "binary_root_spaced": bi_root_spaced,
            "letter1": letter1,
            "letter1_meaning": letter1_meaning,
            "letter2": letter2,
            "letter2_meaning": letter2_meaning,
            "binary_root_meaning": bi_meaning,
            "tri_root": tri_root,
            "tri_root_spaced": tri_root_spaced,
            "added_letter": added_letter,
            "axial_meaning": axial_meaning,
            "quran_example": quran_example,
        })

    wb.close()
    return records


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    sys.stdout.reconfigure(encoding="utf-8")

    # Resolve paths relative to this script's location
    script_dir  = Path(__file__).parent
    lv1_root    = script_dir.parent
    muajam_dir  = lv1_root.parent / "Muajam Ishtiqaqi"
    tables_dir  = muajam_dir / "Tables_Juthoor"
    master_xlsx = muajam_dir / "المعجم_الاشتقاقي_Juthoor_v2.xlsx"

    out_dir = lv1_root / "data" / "muajam"
    out_dir.mkdir(parents=True, exist_ok=True)

    letter_meanings_path = out_dir / "letter_meanings.jsonl"
    roots_path           = out_dir / "roots.jsonl"

    # --- Letter meanings ---
    print(f"Reading master XLSX: {master_xlsx}")
    letter_meanings = extract_letter_meanings(master_xlsx)
    print(f"  Found {len(letter_meanings)} letter meanings")

    with open(letter_meanings_path, "w", encoding="utf-8") as f:
        for rec in letter_meanings:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"  Written: {letter_meanings_path}")

    # --- Roots from per-letter XLSX files ---
    xlsx_files = sorted(tables_dir.glob("*.xlsx"))
    print(f"\nFound {len(xlsx_files)} per-letter XLSX files in {tables_dir}")

    all_roots = []
    for xlsx_path in xlsx_files:
        try:
            roots = extract_roots_from_file(xlsx_path)
            all_roots.extend(roots)
            print(f"  {xlsx_path.name}: {len(roots)} entries")
        except Exception as exc:
            print(f"  ERROR in {xlsx_path.name}: {exc}")

    print(f"\nTotal root entries: {len(all_roots)}")

    with open(roots_path, "w", encoding="utf-8") as f:
        for rec in all_roots:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")
    print(f"Written: {roots_path}")


if __name__ == "__main__":
    main()
